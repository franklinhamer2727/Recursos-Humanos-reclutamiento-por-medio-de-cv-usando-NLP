from gensim.models import Word2Vec
import openpyxl

ruta_modelo_experiencia_laboral = '../Modelos/experiencia_laboral.model'
ruta_modelo_habilidades = '../Modelos/habilidades.model'
ruta_excel = "../RecursosHumanos/rrhh.xlsx"

def lectura_excel(ruta_excel):
    wb = openpyxl.load_workbook(ruta_excel)
    ws_habilidades = wb['habilidades']
    ws_experiencia = wb['experiencia']
    values_experiencia = [ws_experiencia.cell(row=i, column=1).value for i in range(2, ws_experiencia.max_row + 1) if ws_experiencia.cell(row=i, column=1).value is not None and ws_experiencia.cell(row=i, column=1).value != '']
    values_habilidades = [ws_habilidades.cell(row=i, column=1).value for i in range(2, ws_habilidades.max_row + 1) if ws_habilidades.cell(row=i, column=1).value is not None and ws_habilidades.cell(row=i, column=1).value != '']
    return values_experiencia, values_habilidades

def lectura_modelo(ruta, nombre_modelo) -> Word2Vec:
    modelo = None
    if nombre_modelo == "habilidades":
        print("modelo de habilidades")
        modelo = Word2Vec.load(ruta)
    elif nombre_modelo == "experiencia_laboral":
        print("modelo de experiencia_laboral")
        modelo = Word2Vec.load(ruta)
    return modelo

def aplicar_modelo(arr_values, modelo):
    palabras_cercanas = None
    array_respuestas = []
    for i in arr_values:
        print(i)
        if i in modelo.wv.key_to_index:  # Verificar si la palabra está en el vocabulario
            palabras_cercanas = modelo.wv.most_similar(i, topn=10)
            array_respuestas.append(palabras_cercanas)
        else:
            pass
            #print(f"La palabra '{i}' no está en el vocabulario del modelo.")
    return array_respuestas



def procesos_ejecucion():
    values_experiencia, values_habilidades = lectura_excel(ruta_excel)

    modelo_experiencia_laboral = lectura_modelo(ruta_modelo_experiencia_laboral, 'experiencia_laboral')
    palabras_cercanas_experiencia = aplicar_modelo(values_experiencia, modelo_experiencia_laboral)

    modelo_habilidades = lectura_modelo(ruta_modelo_habilidades, 'habilidades')
    palabras_cercanas_habilidades = aplicar_modelo(values_habilidades, modelo_habilidades)
    return  palabras_cercanas_experiencia, palabras_cercanas_habilidades
if __name__ == '__main__':
    palabras_cercanas_experiencia, palabras_cercanas_habilidades = procesos_ejecucion()
    print(palabras_cercanas_experiencia)



